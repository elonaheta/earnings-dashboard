import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import os




# CONFIG

st.set_page_config(
    page_title="Earnings Dashboard",
    layout="wide",
    initial_sidebar_state="expanded"
)


# CSS



st.markdown("""
<style>
[data-testid="stSidebar"] button { display: none !important; }
[data-testid="stSidebar"] [data-testid="collapsedControl"] { display: none !important; }
section[data-testid="stSidebar"] > div { pointer-events: none; }
section[data-testid="stSidebar"] * { pointer-events: auto; }

section[data-testid="stSidebar"] { min-width: 280px !important; max-width: 280px !important; }

.kpi-card {
    background: linear-gradient(135deg, #111827, #1f2937);
    padding: 18px;
    border-radius: 14px;
    color: white;
    height: 135px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    box-shadow: 0 8px 24px rgba(0,0,0,.15);
}
.kpi-title { font-size: 14px; opacity: .8; }
.kpi-value { font-size: 28px; font-weight: 700; }
.kpi-sub { font-size: 13px; opacity: .7; }

h1,h2,h3 { font-weight: 700; }

.block-container { padding-top: 2rem; }
.stTabs [data-baseweb="tab-list"] { gap: 1rem; }
.stPlotlyChart { border-radius: 8px; }

.section-header {
    margin: 2rem 0 1rem 0;
    padding-bottom: 0.5rem;
    border-bottom: 1px solid #e5e7eb;
}
</style>
""", unsafe_allow_html=True)

# DATA LOADING 

@st.cache_data
def load_excel(file_path="data/Data.App.xlsx"):
    if not os.path.exists(file_path):
        st.error(f"Excel file not found at: {file_path}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    daily, conductors, locations = [], [], []
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        date = ws["E1"].value
        if not date:
            continue
        date = pd.to_datetime(date)
        
        people = ws["B1"].value or 0
        earnings = ws["G13"].value or 0
        parking = ws["H13"].value or 0
        net = ws["I13"].value or (earnings - parking)
        
        daily.append({
            "date": date,
            "people": people,
            "earnings": earnings,
            "parking": parking,
            "net": net,
            "dow": date.strftime("%A")
        })
        
        for r in range(5, 15):
            conductor = ws[f"B{r}"].value
            location = ws[f"C{r}"].value
            if not conductor:
                continue
            e = ws[f"G{r}"].value or 0
            p = ws[f"H{r}"].value or 0
            n = ws[f"I{r}"].value or (e - p)
            
            conductors.append({
                "date": date,
                "conductor": conductor,
                "location": location or "Unknown",
                "earnings": e,
                "parking": p,
                "net": n
            })
            
            locations.append({
                "date": date,
                "location": location or "Unknown",
                "earnings": e,
                "net": n
            })
    
    return pd.DataFrame(daily), pd.DataFrame(conductors), pd.DataFrame(locations)

daily_df, conductor_df, location_df = load_excel()

if daily_df.empty:
    st.stop()  




# SIDEBAR 

with st.sidebar:
    st.title("Filters")
    st.markdown("---")
    
    min_d = daily_df.date.min().date()
    max_d = daily_df.date.max().date()
    
    start, end = st.date_input(
        "Date Range",
        (min_d, max_d),
        min_value=min_d,
        max_value=max_d
    )

# Filter data
daily_df = daily_df[(daily_df.date.dt.date >= start) & (daily_df.date.dt.date <= end)]
conductor_df = conductor_df[(conductor_df.date.dt.date >= start) & (conductor_df.date.dt.date <= end)]
location_df = location_df[(location_df.date.dt.date >= start) & (location_df.date.dt.date <= end)]






# HEADER


st.title("Earnings Dashboard")
st.caption(f"Period: {start} to {end}")


# KPI 

def kpi(col, title, total, avg, suffix=""):
    col.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-title">{title}</div>
        <div class="kpi-value">{total}{suffix}</div>
        <div class="kpi-sub">Avg/day: {avg}{suffix}</div>
    </div>
    """, unsafe_allow_html=True)

# KPI ROW
st.markdown('<div class="section-header">Performance Summary</div>', unsafe_allow_html=True)
c1, c2, c3, c4 = st.columns(4)

kpi(c1, "Total Net", f"{daily_df.net.sum():,.0f}", f"{daily_df.net.mean():,.0f}", " €")
kpi(c2, "Total Earnings", f"{daily_df.earnings.sum():,.0f}", f"{daily_df.earnings.mean():,.0f}", " €")
kpi(c3, "Total Parking", f"{daily_df.parking.sum():,.0f}", f"{daily_df.parking.mean():,.0f}", " €")
kpi(c4, "Total Passengers", f"{daily_df.people.sum():,.0f}", f"{int(daily_df.people.mean()):,}")




# DAILY TRENDS

st.markdown('<div class="section-header">Daily Trends</div>', unsafe_allow_html=True)

def plot_daily_trends(df):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df.date, y=df.net, name="Net", mode="lines+markers", line=dict(width=3)))
    fig.add_trace(go.Scatter(x=df.date, y=df.earnings, name="Earnings", line=dict(width=2, dash='dot')))
    fig.add_trace(go.Scatter(x=df.date, y=df.parking, name="Parking", line=dict(width=2)))
    
    fig.update_layout(
        height=400,
        template="plotly_white",
        hovermode="x unified",
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=0, r=0, t=30, b=0)
    )
    return fig

st.plotly_chart(plot_daily_trends(daily_df), use_container_width=True)




# TABS: Locations, Conductors, Weekly Analysis
 

tab1, tab2, tab3 = st.tabs(["Locations", "Conductors", "Weekly Analysis"])


# --- Locations ---
with tab1:
    st.markdown("### Location Performance")
    loc_sum = location_df.groupby("location").agg(net=("net", "sum"), earnings=("earnings", "sum")).reset_index().sort_values("net", ascending=True)
    col1, col2 = st.columns([2,1])
    
    with col1:
        fig = px.bar(loc_sum.tail(10), y="location", x="net", orientation='h', title="Net Earnings by Location", color="net", color_continuous_scale="blues")
        fig.update_layout(height=400, showlegend=False)
        st.plotly_chart(fig, use_container_width=True)
    
    with col2:
        fig = px.pie(loc_sum, values="earnings", names="location", hole=0.4, title="Earnings Distribution")
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)


# --- Conductors ---
with tab2:
    st.markdown("### Conductor Performance")
    cond_sum = conductor_df.groupby("conductor").agg(net=("net","sum"), earnings=("earnings","sum"), parking=("parking","sum")).reset_index().sort_values("net", ascending=True)
    fig = px.bar(cond_sum.tail(10), y="conductor", x="net", orientation='h', title="Top Conductors by Net Earnings", color="net", color_continuous_scale="greens")
    fig.update_layout(height=500)
    st.plotly_chart(fig, use_container_width=True)


# --- Weekly Analysis ---
with tab3:
    st.markdown("### Weekly Analysis")
    dow_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    dow = daily_df.groupby("dow").agg(net=("net","mean"), people=("people","mean")).reindex(dow_order).fillna(0)
    
    col1, col2 = st.columns(2)
    
    with col1:
        fig = px.bar(dow, x=dow.index, y="net", title="Average Net by Day", color="net", color_continuous_scale="purples")
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
    
    with col2:
        fig = px.bar(dow, x=dow.index, y="people", title="Average Passengers by Day", color="people", color_continuous_scale="oranges")
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)



# DATA TABLES


with st.expander("View Data Tables"):
    tab1, tab2, tab3 = st.tabs(["Daily Summary", "Conductor Details", "Location Details"])
    
    with tab1:
        st.dataframe(daily_df.sort_values("date", ascending=False).reset_index(drop=True), use_container_width=True)
    with tab2:
        st.dataframe(conductor_df.sort_values(["date","conductor"]).reset_index(drop=True), use_container_width=True)
    with tab3:
        st.dataframe(location_df.sort_values(["date","location"]).reset_index(drop=True), use_container_width=True)



# FOOTER



st.markdown("---")
st.caption(f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")


